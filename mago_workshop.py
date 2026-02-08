import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from datetime import datetime

# Configurazione percorsi
BASE = os.getcwd()
SRC = os.path.join(BASE, "src")
DOCS = os.path.join(BASE, "docs")
SCRIPTS = os.path.join(BASE, "scripts")
BACKUPS = os.path.join(BASE, "backups")

def esegui_mago():
    print("--- Avvio Trasformazione Totale ---")
    
    # 1. Creazione Cartelle
    for d in [SRC, DOCS, SCRIPTS, BACKUPS]:
        if not os.path.exists(d): os.makedirs(d)

    # 2. Spostamento file esistenti
    for f in os.listdir(BASE):
        if f.endswith(".txt") and f[0].isdigit():
            shutil.move(f, os.path.join(DOCS, f))
        elif f.endswith(".xlsm") and "BACKUP" not in f:
            shutil.move(f, os.path.join(SRC, f))

    # 3. Creazione Dashboard interna all'Excel
    excel_path = os.path.join(SRC, "CARICO_LAVORO_OFFICINA_Rev0.xlsm")
    if os.path.exists(excel_path):
        print("[...] Creazione Dashboard nell'Excel...")
        try:
            df = pd.read_excel(excel_path, sheet_name='ELENCO_COMMESSE').dropna(subset=['COMMESSA'])
            wb = openpyxl.load_workbook(excel_path, keep_vba=True)
            if 'DASHBOARD' in wb.sheetnames: del wb['DASHBOARD']
            ws = wb.create_sheet('DASHBOARD', 0)
            
            # Grafica
            ws['A1'] = "RIEPILOGO PRODUZIONE OFFICINA"
            ws['A1'].font = Font(name='Tahoma', size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws['A3'], ws['B3'] = "TOTALE COMMESSE:", len(df)
            ws['A4'], ws['B4'] = "COMPLETATE:", df['Fine lavorazione TOTALE'].notna().sum()
            ws['A3'].font = ws['A4'].font = Font(bold=True)
            
            # Carico Reparti
            ws['D3'] = "ORE PER REPARTO"
            ws['D3'].font = Font(bold=True)
            fasi = [c for c in df.columns if str(c).startswith('ORE ')]
            carico = df[fasi].sum()
            r = 4
            for f, o in carico.items():
                ws.cell(row=r, column=4).value = str(f).replace('ORE ', '')
                ws.cell(row=r, column=5).value = o
                r += 1
            
            wb.save(excel_path)
            print("[OK] Dashboard creata con successo!")
        except Exception as e:
            print(f"[ERRORE] Dashboard: {e}")

    # 4. Creazione script Backup corretti
    with open(os.path.join(SCRIPTS, "backup_manager.py"), "w") as f:
        f.write("import shutil, os\\nfrom datetime import datetime\\ndef create_backup():\\n    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))\\n    src = os.path.join(base, 'src', 'CARICO_LAVORO_OFFICINA_Rev0.xlsm')\\n    dst = os.path.join(base, 'backups')\\n    if not os.path.exists(dst): os.makedirs(dst)\\n    ts = datetime.now().strftime('%Y%m%d_%H%M%S')\\n    shutil.copy2(src, os.path.join(dst, f'BACKUP_{ts}.xlsm'))\\n    print('Backup creato.')\\nif __name__ == '__main__': create_backup()")

    print("\\n[FINITO!] Ora il tuo PC è aggiornato.")
    print("Per aggiornare GitHub, scrivi: git add . && git commit -m 'Dashboard e rami' && git push origin main")

if __name__ == "__main__":
    esegui_mago()